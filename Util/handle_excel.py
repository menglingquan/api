import openpyxl
import os
base_path = os.path.abspath(os.path.dirname(os.getcwd()))

class HandleExcel:
    def load_excel(self):
        '''
        加载excel
        '''
        open_excel = openpyxl.load_workbook(base_path+"\Case\First.xlsx")
        return open_excel
    def get_sheet_data(self,index=None):
        '''
        加载所有sheet内容
        '''
        sheet_name = self.load_excel().sheetnames
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data
    def get_cell_value(self,row,cols):
        '''
        获取单元格内容
        '''
        data = self.get_sheet_data().cell(row=row,column=cols).value
        return data
    def get_rows(self):
        row = self.get_sheet_data().max_row
        return row
    def get_rows_value(self,row):
        #获取某一行内容
        rows_list = []
        for i in self.get_sheet_data()[row]:
            rows_list.append(i.value)
        return rows_list
excel_data =HandleExcel()
if __name__ == '__main__':
    handle = HandleExcel()
    print(handle.get_cell_value(2,1))
    print(handle.get_rows_value(2))